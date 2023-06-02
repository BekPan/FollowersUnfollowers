from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

book = Workbook()

sheet = book.active
sheet.title = "Data"
sheet["A1"] = "Followers"
sheet["B1"] = "Following"
sheet["A1"].font = Font(bold=True)
sheet["B1"].font = Font(bold=True)
sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 40


fileName = "user.txt"
delimeter = '='
file = open(fileName, 'r')

def findValue(fullString):
    fullString = fullString.rstrip('\n')
    value = fullString[fullString.index(delimeter)+1:]
    value = value.replace(" ","")
    return value
for line in file:
    if line.startswith('username'):
        username = findValue(line)
    if line.startswith('password'):
        password = findValue(line)
    if line.startswith('followers'):
        followers = findValue(line)
    if line.startswith('following'):
        following = findValue(line)
file.close()
from playwright.sync_api import sync_playwright

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False, slow_mo=250)
    page =  browser.new_page()
    page.goto('https://www.instagram.com/accounts/login/?hl=en')
    page.get_by_role("button", name="Decline optional cookies").click()
    page.locator('//*[@id="loginForm"]/div/div[1]/div/label/input').fill(username)
    page.locator('//*[@id="loginForm"]/div/div[2]/div/label/input').fill(password)
    page.locator('//*[@id="loginForm"]/div/div[3]/button/div').click()
    page.get_by_role("button", name="Not Now").click()
    page.get_by_role("button", name="Not Now").click()
    user=username+"'s profile picture Profile"
    page.get_by_role("link", name=user).click()
    page.get_by_role("link", name="followers").click()
    
    #playwright.$("div._aano div div span a >> nth=0") for a specific element
    #playwright.$$("div._aano div div span a") for all elements
    f1 = 0
    row1 = 2
    while f1 < int(followers) :
        if f1 == 0:
            print("===Followers===")
        follower = page.locator('div._aano div div span a >> nth='+str(f1)).inner_text()
        followerSCR = page.locator('div._aano div div span a >> nth='+str(f1))
        followerSCR.scroll_into_view_if_needed()
        print(follower)
        sheet.cell(row=row1, column=1, value=follower)
        row1 += 1
        f1 += 1
    page.get_by_role("button", name="Close").click()

    page.get_by_role("link", name="following").click()
    f2 = 0
    row2 = 2
    while f2 < int(following) :
        if f2 == 0:
            print("===Following===")
        followi = page.locator('div._aano div div span a >> nth='+str(f2)).inner_text()
        followiSCR = page.locator('div._aano div div span a >> nth='+str(f2))
        followiSCR.scroll_into_view_if_needed()
        print(followi)
        sheet.cell(row=row2, column=2, value=followi)
        row2 += 1
        f2 += 1
    page.get_by_role("button", name="Close").click()

    page.get_by_role("link", name="Settings More").click()
    page.get_by_role("button", name="Log out").click()

    sheet["C1"] = "Not Following"
    sheet["D1"] = "Not Followers"
    sheet["C1"].font = Font(bold=True)
    sheet["D1"].font = Font(bold=True)
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40

    Aend = int(followers) + 1
    Bend = int(following) + 1
    C2 = "FILTER(A2:A" + str(Aend) + ",COUNTIF(B2:B" + str(Bend) + ",A2:A" + str(Aend) + ")=0)"
    D2 = "FILTER(B2:B" + str(Bend) + ",COUNTIF(A2:A" + str(Aend) + ",B2:B" + str(Bend) + ")=0)"
    sheet.cell(row=2, column=3, value=C2)
    sheet.cell(row=2, column=4, value=D2)
    sheet["E2"] = "In cells C2,D2 with formula FILTER put = at the front to load results"
    sheet["E2"].font = Font(bold=True)
    sheet.column_dimensions['E'].width = 60
    print("In cells C2,D2 with formula FILTER put = at the front to load results")

    book.save("followersUnfollowers.xlsx")